"""结果打包服务 — xlsx 生成 + zip 打包。

提供:
  - build_summary_xlsx(): 从求解报告提取数据表，生成多 sheet xlsx
  - build_zip_package(): 打包论文 + 数据 + 图表 + 代码为 zip
"""

import logging
import re
import zipfile
from pathlib import Path

logger = logging.getLogger(__name__)


class ResultPackager:
    """收集求解产出的所有文件，生成汇总 xlsx 和 zip 打包下载。"""

    def __init__(self, task_id: str, project_root: Path):
        self.task_id = task_id
        self.project_root = project_root
        self.task_dir = project_root / "data" / "task_files" / task_id

    # ── xlsx 生成 ─────────────────────────────────────────────────

    def build_summary_xlsx(
        self,
        solving_output: str = "",
        task_files_dir: Path | None = None,
    ) -> Path:
        """从求解报告中解析数据表，生成汇总 xlsx。

        包含 sheet:
          - "最优解"：决策变量最优值
          - "参数扫描"：灵敏度分析参数扫描表
          - "结果汇总"：各子问题关键结果

        如果 LLM 已经生成了 .xlsx 文件，直接返回第一个。
        """
        # 优先使用 LLM 已生成的 xlsx
        if task_files_dir and task_files_dir.exists():
            existing_xlsx = sorted(task_files_dir.glob("*.xlsx"))
            if existing_xlsx:
                return existing_xlsx[0]

        # 否则从 Markdown 中提取表格生成 xlsx
        output_path = self.task_dir / "results_summary.xlsx"
        self.task_dir.mkdir(parents=True, exist_ok=True)

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.warning("pandas/openpyxl 未安装，跳过 xlsx 生成")
            return output_path

        wb = Workbook()

        # 从求解报告中解析 Markdown 表格
        tables = self._extract_markdown_tables(solving_output)

        if tables:
            for sheet_name, rows in tables.items():
                if sheet_name == "最优解":
                    ws = wb.active
                    ws.title = "最优解"
                else:
                    ws = wb.create_sheet(title=sheet_name[:31])  # sheet 名限 31 字符
                try:
                    df = pd.DataFrame(rows[1:], columns=rows[0])
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                except Exception as e:
                    ws.cell(row=1, column=1, value=f"解析失败: {e}")
        else:
            # 兜底：插入求解报告全文为单个 sheet
            ws = wb.active
            ws.title = "求解报告"
            ws.cell(row=1, column=1, value="求解计算输出")
            for i, line in enumerate(solving_output.split("\n")[:500], 2):
                ws.cell(row=i, column=1, value=line[:200])

        try:
            wb.save(output_path)
        except Exception as e:
            logger.warning("xlsx 保存失败: %s", e)

        return output_path

    def _extract_markdown_tables(self, text: str) -> dict[str, list[list[str]]]:
        """从 Markdown 文本中提取表格，按上下文分组。

        返回 {sheet_name: [header_row, ...data_rows]}。
        """
        if not text:
            return {}

        tables: dict[str, list[list[str]]] = {}
        current_sheet = "求解结果"
        current_table: list[list[str]] = []
        in_table = False

        for line in text.split("\n"):
            stripped = line.strip()

            # 检测小节标题 → 作为 sheet 名
            if stripped.startswith("### ") or stripped.startswith("## "):
                title = stripped.lstrip("#").strip()[:30]
                if "最优解" in title:
                    current_sheet = "最优解"
                elif any(kw in title for kw in ("参数扫描", "灵敏度", "sensitivity")):
                    current_sheet = "参数扫描"
                elif any(kw in title for kw in ("结果汇总", "求解小结", "summary")):
                    current_sheet = "结果汇总"
                # 如果已有表格数据，先保存
                if current_table and len(current_table) >= 2:
                    tables.setdefault(current_sheet, current_table)
                    current_table = []
                continue

            # 检测 Markdown 表格行
            if "|" in stripped and stripped.startswith("|"):
                cells = [c.strip() for c in stripped.split("|")[1:-1]]
                # 跳过分隔行 (|---|---|)
                if all(re.match(r"^:?-{2,}:?$", c) for c in cells if c):
                    continue
                current_table.append(cells)
                in_table = True
            elif in_table and not stripped:
                # 空行结束表格
                if current_table and len(current_table) >= 2:
                    tables.setdefault(current_sheet, current_table)
                    current_table = []
                in_table = False

        # 保存最后一张表
        if current_table and len(current_table) >= 2:
            tables.setdefault(current_sheet, current_table)

        return tables

    # ── zip 打包 ─────────────────────────────────────────────────

    def build_zip_package(self) -> Path:
        """打包任务产出的所有文件。

        包含:
          - 论文.md（final_response）
          - results.xlsx（求解结果表）
          - figures/（所有图表 PNG）
          - data/（用户上传的原始数据文件）
          - code/（提取的求解代码片段）

        返回 zip 文件路径。
        """
        output_path = self.task_dir / f"{self.task_id}_results.zip"
        self.task_dir.mkdir(parents=True, exist_ok=True)

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in sorted(self.task_dir.iterdir()):
                if f.is_file() and f.suffix.lower() in (".png", ".xlsx", ".csv", ".html"):
                    arcname = f.name
                    # 图表放到 figures/ 下
                    if f.suffix.lower() == ".png":
                        arcname = f"figures/{f.name}"
                    elif f.suffix.lower() == ".xlsx":
                        arcname = f"data/{f.name}"
                    elif f.suffix.lower() == ".csv":
                        arcname = f"data/{f.name}"
                    elif f.suffix.lower() == ".html":
                        arcname = f"reports/{f.name}"
                    zf.write(f, arcname=arcname)

                    # 注册到 session manager
                    from app.services.session import get_session_manager

                    get_session_manager().add_artifact(
                        self.task_id,
                        {
                            "type": f.suffix.lstrip("."),
                            "name": f.name,
                            "url": f"/api/task_files/{self.task_id}/{f.name}",
                            "size": f.stat().st_size,
                        },
                    )

        # 注册 zip 到 session manager
        from app.services.session import get_session_manager

        get_session_manager().add_artifact(
            self.task_id,
            {
                "type": "zip",
                "name": f"{self.task_id}_results.zip",
                "url": f"/api/task_files/{self.task_id}/{self.task_id}_results.zip",
                "size": output_path.stat().st_size,
            },
        )

        return output_path
